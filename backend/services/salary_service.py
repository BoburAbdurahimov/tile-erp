import calendar
import io
import logging
from datetime import date, datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func

from backend.models import (
    Employee, JobType, AttendanceEntry, WorkEntry,
    MonthlySalaryCalculation, CashRegister, CashTransaction, AuditLog
)

logger = logging.getLogger(__name__)

def get_month_date_range(year_month: str):
    """Parse 'YYYY-MM' and return (start_date, end_date, total_days_in_month)."""
    parts = year_month.split("-")
    year = int(parts[0])
    month = int(parts[1])
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)
    return start_date, end_date, num_days

def calculate_employee_salary(
    db: Session,
    employee_id: int,
    year_month: str,
    bonus: float = 0.0,
    advance: float = 0.0,
    notes: Optional[str] = None,
    current_user: str = "Admin"
) -> MonthlySalaryCalculation:
    """Calculate and upsert salary calculation for a specific employee and month."""
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise ValueError("Xodim topilmadi (Employee not found)")

    start_date, end_date, _ = get_month_date_range(year_month)

    # Check if existing calculation is locked/paid
    calc = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.employee_id == employee_id,
        MonthlySalaryCalculation.year_month == year_month
    ).first()

    if calc and calc.status in ("finalized", "paid"):
        return calc

    if not calc:
        calc = MonthlySalaryCalculation(
            employee_id=employee_id,
            year_month=year_month,
            employee_type=employee.employee_type,
            status="draft"
        )
        db.add(calc)

    calc.employee_type = employee.employee_type
    if bonus > 0:
        calc.bonus_amount = bonus
    if advance > 0:
        calc.advance_paid = advance
    if notes is not None:
        calc.notes = notes

    if employee.employee_type == "fixed":
        base_salary = float(employee.monthly_salary or 0.0)
        standard_days = int(employee.standard_work_days or 26)
        
        # Proration for mid-month hire or removal
        if employee.hire_date and employee.hire_date > start_date:
            # Hired mid-month: adjust standard days proportionally
            days_active = (end_date - employee.hire_date).days + 1
            effective_days = max(1, min(standard_days, int(round((days_active / 30.0) * standard_days))))
        else:
            effective_days = standard_days

        # Count absent days in this month
        absent_entries = db.query(AttendanceEntry).filter(
            AttendanceEntry.employee_id == employee_id,
            AttendanceEntry.date >= start_date,
            AttendanceEntry.date <= end_date,
            AttendanceEntry.status == "absent"
        ).all()
        absent_days = len(absent_entries)

        per_day_rate = round(base_salary / max(1, effective_days), 2)
        deduction_amount = min(base_salary, round(per_day_rate * absent_days, 2))
        final_amount = max(0.0, round(base_salary - deduction_amount + (calc.bonus_amount or 0.0) - (calc.advance_paid or 0.0), 2))

        calc.base_salary = base_salary
        calc.standard_days = effective_days
        calc.absent_days = absent_days
        calc.per_day_rate = per_day_rate
        calc.deduction_amount = deduction_amount
        calc.piecework_total = 0.0
        calc.final_amount = final_amount

    elif employee.employee_type == "piecework":
        # Sum of work entries
        work_entries = db.query(WorkEntry).filter(
            WorkEntry.employee_id == employee_id,
            WorkEntry.date >= start_date,
            WorkEntry.date <= end_date
        ).all()
        
        piecework_total = round(sum(float(w.total_amount or 0.0) for w in work_entries), 2)
        final_amount = max(0.0, round(piecework_total + (calc.bonus_amount or 0.0) - (calc.advance_paid or 0.0), 2))

        calc.base_salary = 0.0
        calc.standard_days = 0
        calc.absent_days = 0
        calc.per_day_rate = 0.0
        calc.deduction_amount = 0.0
        calc.piecework_total = piecework_total
        calc.final_amount = final_amount

    calc.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(calc)
    return calc

def recalculate_all_salaries(db: Session, year_month: str, current_user: str = "Admin") -> List[MonthlySalaryCalculation]:
    """Calculate/refresh salaries for all active employees for given month."""
    start_date, end_date, _ = get_month_date_range(year_month)
    
    # Active employees or employees who worked during this month
    employees = db.query(Employee).filter(
        (Employee.is_active == True) | 
        ((Employee.removal_date != None) & (Employee.removal_date >= start_date))
    ).all()

    results = []
    for emp in employees:
        # Ignore employees hired after this month
        if emp.hire_date and emp.hire_date > end_date:
            continue
        calc = calculate_employee_salary(db, emp.id, year_month, current_user=current_user)
        results.append(calc)
    return results

def get_payroll_summary(db: Session, year_month: str) -> Dict[str, Any]:
    """Get full calculation report with summary statistics and details."""
    calcs = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.year_month == year_month
    ).join(Employee).order_by(Employee.employee_type.asc(), Employee.full_name.asc()).all()

    total_payroll = sum(c.final_amount for c in calcs)
    total_fixed = sum(c.final_amount for c in calcs if c.employee_type == "fixed")
    total_piecework = sum(c.final_amount for c in calcs if c.employee_type == "piecework")
    total_paid = sum(c.final_amount for c in calcs if c.status == "paid")
    total_unpaid = total_payroll - total_paid

    items = []
    for c in calcs:
        emp = c.employee
        items.append({
            "id": c.id,
            "employee_id": emp.id,
            "full_name": emp.full_name,
            "department": emp.department or "Ma'muriyat",
            "employee_type": emp.employee_type,
            "position": emp.position or "-",
            "phone_number": emp.phone_number or "-",
            "base_salary": c.base_salary,
            "standard_days": c.standard_days,
            "absent_days": c.absent_days,
            "per_day_rate": c.per_day_rate,
            "deduction_amount": c.deduction_amount,
            "piecework_total": c.piecework_total,
            "bonus_amount": c.bonus_amount or 0.0,
            "advance_paid": c.advance_paid or 0.0,
            "final_amount": c.final_amount,
            "status": c.status,
            "cash_transaction_id": c.cash_transaction_id,
            "finalized_at": c.finalized_at.isoformat() if c.finalized_at else None,
            "paid_at": c.paid_at.isoformat() if c.paid_at else None,
            "paid_by": c.paid_by,
            "notes": c.notes
        })

    is_all_finalized = len(calcs) > 0 and all(c.status in ("finalized", "paid") for c in calcs)

    return {
        "year_month": year_month,
        "is_all_finalized": is_all_finalized,
        "total_employees": len(calcs),
        "total_payroll": round(total_payroll, 2),
        "total_fixed": round(total_fixed, 2),
        "total_piecework": round(total_piecework, 2),
        "total_paid": round(total_paid, 2),
        "total_unpaid": round(total_unpaid, 2),
        "calculations": items
    }

def record_daily_attendance(
    db: Session,
    entry_date: date,
    absent_records: List[Dict[str, Any]],
    current_user: str = "Admin"
):
    """Save attendance entries for a date and update monthly salary."""
    year_month = entry_date.strftime("%Y-%m")
    
    # Check if month is finalized
    locked = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.year_month == year_month,
        MonthlySalaryCalculation.status.in_(["finalized", "paid"])
    ).first()
    if locked:
        raise ValueError(f"{year_month} oyi allaqachon yopilgan va qulflangan!")

    absent_emp_ids = {r["employee_id"] for r in absent_records}

    # Delete existing entries for this date for fixed employees
    existing = db.query(AttendanceEntry).filter(AttendanceEntry.date == entry_date).all()
    for e in existing:
        if e.employee_id not in absent_emp_ids:
            db.delete(e)

    # Insert or update absent entries
    for r in absent_records:
        emp_id = r["employee_id"]
        reason = r.get("reason", "")
        att = db.query(AttendanceEntry).filter(
            AttendanceEntry.employee_id == emp_id,
            AttendanceEntry.date == entry_date
        ).first()
        if not att:
            att = AttendanceEntry(
                employee_id=emp_id,
                date=entry_date,
                status="absent",
                reason=reason,
                entered_by=current_user
            )
            db.add(att)
        else:
            att.status = "absent"
            att.reason = reason
            att.entered_by = current_user

    db.commit()

    db.commit()

    # Recalculate salaries for affected fixed employees
    fixed_employees = db.query(Employee).filter(
        Employee.employee_type == "fixed",
        Employee.is_active == True
    ).all()
    for emp in fixed_employees:
        calculate_employee_salary(db, emp.id, year_month, current_user=current_user)

    # Sync piecework earnings for present workers on all 5 production lines based on actual daily production volume
    from backend.models import ProductionLine
    lines = db.query(ProductionLine).all()
    for l in lines:
        sync_piecework_from_production(db, l.id, entry_date)

    db.add(AuditLog(
        username=current_user,
        action="UPDATE",
        module="Ish haqi / Davomat",
        entity_id=str(entry_date),
        details=f"{entry_date} sanasi uchun {len(absent_records)} ta kelmagan xodim davomati saqlandi"
    ))
    db.commit()
    return True

def sync_piecework_from_production(db: Session, line_id: int, date_val: date, total_output_qty: float = None):
    """
    Automatically syncs piecework earnings for present workers on a production line and date
    based on the actual production volume output (e.g. 170-meter kiln run / tile output).
    """
    from backend.models import ProductionLine, ProductionOrder
    
    line = db.query(ProductionLine).filter(ProductionLine.id == line_id).first()
    if not line:
        return

    if total_output_qty is None:
        total_output_qty = db.query(func.sum(ProductionOrder.quantity)).filter(
            ProductionOrder.line_id == line_id,
            ProductionOrder.date == date_val,
            ProductionOrder.status == "Tasdiqlandi"
        ).scalar() or 0.0

    dept_name = f"{line.line_number}-Liniya"
    
    piecework_emps = db.query(Employee).filter(
        Employee.department == dept_name,
        Employee.employee_type == "piecework",
        Employee.is_active == True
    ).all()

    if not piecework_emps:
        return

    absent_emp_ids = set(
        a.employee_id for a in db.query(AttendanceEntry).filter(
            AttendanceEntry.date == date_val,
            AttendanceEntry.status == "absent",
            AttendanceEntry.employee_id.in_([e.id for e in piecework_emps])
        ).all()
    )

    job_types = db.query(JobType).filter(JobType.is_active == True).all()
    jt_map = {jt.name: jt for jt in job_types}

    year_month = date_val.strftime("%Y-%m")

    for emp in piecework_emps:
        if emp.id in absent_emp_ids or total_output_qty <= 0:
            # Absent worker or 0 production output -> delete work entries on this date
            db.query(WorkEntry).filter(
                WorkEntry.employee_id == emp.id,
                WorkEntry.date == date_val
            ).delete()
            calculate_employee_salary(db, emp.id, year_month)
            continue

        # Worker is PRESENT and line produced output!
        target_jt = None
        qty = total_output_qty

        pos = (emp.position or "").lower()
        if "press" in pos:
            tile_size = (line.spec_tile_size or "30x30").split()[0]
            target_jt = jt_map.get(f"Pressovka va Formovka ({tile_size})") or jt_map.get("Pressovka va Formovka (30x30)")
        elif "glazur" in pos:
            target_jt = jt_map.get("Glazurlash va Linya Bo'yoq")
            qty = total_output_qty * 0.95
        elif "pech" in pos:
            target_jt = jt_map.get("Pechda Kuydirish Nazorati")
            qty = total_output_qty * 0.92
        elif "saral" in pos:
            target_jt = jt_map.get("Saralash va Sifat Nazorati")
            qty = total_output_qty * 0.90
        else: # Qadoqlovchi / Poddon
            target_jt = jt_map.get("Qadoqlash va Poddon Yig'ish")
            qty = round(total_output_qty / 40.0, 1)

        if not target_jt:
            continue

        rate = target_jt.price_per_unit
        total_amt = round(qty * rate, 2)

        we = db.query(WorkEntry).filter(
            WorkEntry.employee_id == emp.id,
            WorkEntry.job_type_id == target_jt.id,
            WorkEntry.date == date_val
        ).first()

        note_text = f"Liniya #{line.line_number} ({total_output_qty:,.0f} dona/m2 ishlab chiqarishdan avto-hisoblandi)"

        if not we:
            we = WorkEntry(
                employee_id=emp.id,
                job_type_id=target_jt.id,
                date=date_val,
                quantity=qty,
                unit_price_snapshot=rate,
                total_amount=total_amt,
                notes=note_text,
                entered_by="System (Avto-ishlab chiqarish)"
            )
            db.add(we)
        else:
            we.quantity = qty
            we.unit_price_snapshot = rate
            we.total_amount = total_amt
            we.notes = note_text

        calculate_employee_salary(db, emp.id, year_month)

    db.commit()

def record_daily_work_entry(
    db: Session,
    employee_id: int,
    job_type_id: int,
    entry_date: date,
    quantity: float,
    notes: Optional[str] = None,
    entry_id: Optional[int] = None,
    current_user: str = "Admin"
) -> WorkEntry:
    """Record or update a piecework job entry with price snapshot."""
    year_month = entry_date.strftime("%Y-%m")
    
    # Check if month is finalized
    locked = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.year_month == year_month,
        MonthlySalaryCalculation.status.in_(["finalized", "paid"])
    ).first()
    if locked:
        raise ValueError(f"{year_month} oyi allaqachon yopilgan va qulflangan!")

    job_type = db.query(JobType).filter(JobType.id == job_type_id).first()
    if not job_type:
        raise ValueError("Ish turi topilmadi (JobType not found)")

    unit_price = float(job_type.price_per_unit)
    total_amount = round(quantity * unit_price, 2)

    if entry_id:
        work_entry = db.query(WorkEntry).filter(WorkEntry.id == entry_id).first()
        if not work_entry:
            raise ValueError("Bajarilgan ish yozuvi topilmadi")
        work_entry.employee_id = employee_id
        work_entry.job_type_id = job_type_id
        work_entry.date = entry_date
        work_entry.quantity = quantity
        work_entry.unit_price_snapshot = unit_price
        work_entry.total_amount = total_amount
        work_entry.notes = notes
        work_entry.entered_by = current_user
    else:
        work_entry = WorkEntry(
            employee_id=employee_id,
            job_type_id=job_type_id,
            date=entry_date,
            quantity=quantity,
            unit_price_snapshot=unit_price,
            total_amount=total_amount,
            notes=notes,
            entered_by=current_user
        )
        db.add(work_entry)

    db.commit()
    db.refresh(work_entry)

    # Recalculate salary for this employee
    calculate_employee_salary(db, employee_id, year_month, current_user=current_user)

    db.add(AuditLog(
        username=current_user,
        action="CREATE" if not entry_id else "UPDATE",
        module="Ish haqi / Ishbay",
        entity_id=str(work_entry.id),
        details=f"Ishbay entry: Xodim #{employee_id}, {job_type.name} x {quantity} {job_type.unit_of_measure} = {total_amount:,.0f} UZS"
    ))
    db.commit()
    return work_entry

def delete_daily_work_entry(db: Session, entry_id: int, current_user: str = "Admin") -> bool:
    """Delete a piecework entry and update monthly salary."""
    work_entry = db.query(WorkEntry).filter(WorkEntry.id == entry_id).first()
    if not work_entry:
        raise ValueError("Yozuv topilmadi")

    year_month = work_entry.date.strftime("%Y-%m")
    locked = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.year_month == year_month,
        MonthlySalaryCalculation.status.in_(["finalized", "paid"])
    ).first()
    if locked:
        raise ValueError(f"{year_month} oyi allaqachon yopilgan va qulflangan!")

    emp_id = work_entry.employee_id
    db.delete(work_entry)
    db.commit()

    calculate_employee_salary(db, emp_id, year_month, current_user=current_user)
    
    db.add(AuditLog(
        username=current_user,
        action="DELETE",
        module="Ish haqi / Ishbay",
        entity_id=str(entry_id),
        details=f"Ishbay yozuvi #{entry_id} o'chirildi"
    ))
    db.commit()
    return True

def finalize_month_payroll(db: Session, year_month: str, current_user: str = "Admin"):
    """Lock all calculations for a month."""
    calcs = recalculate_all_salaries(db, year_month, current_user=current_user)
    now = datetime.utcnow()
    for c in calcs:
        if c.status == "draft":
            c.status = "finalized"
            c.finalized_at = now
            c.finalized_by = current_user
    db.commit()

    db.add(AuditLog(
        username=current_user,
        action="FINALIZE",
        module="Ish haqi",
        entity_id=year_month,
        details=f"{year_month} oylik ish haqi hisob-kitobi tasdiqlandi va qulflandi"
    ))
    db.commit()
    return True

def reopen_month_payroll(db: Session, year_month: str, current_user: str = "Admin"):
    """Reopen a finalized month if no salary has been marked as paid."""
    calcs = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.year_month == year_month
    ).all()
    
    paid_count = sum(1 for c in calcs if c.status == "paid")
    if paid_count > 0:
        raise ValueError(f"Ushbu oyda {paid_count} ta xodimga to'lov amalga oshirilgan! Qayta ochish mumkin emas.")

    for c in calcs:
        c.status = "draft"
        c.finalized_at = None
        c.finalized_by = None
    db.commit()

    db.add(AuditLog(
        username=current_user,
        action="REOPEN",
        module="Ish haqi",
        entity_id=year_month,
        details=f"{year_month} oylik ish haqi hisob-kitobi qayta tahrirlash uchun ochildi"
    ))
    db.commit()
    return True

def pay_employee_salary(
    db: Session,
    calculation_id: int,
    register_id: int,
    payment_amount: float,
    current_user: str = "Admin",
    notes: Optional[str] = None
) -> MonthlySalaryCalculation:
    """Pay salary and generate linked cash transaction in Kassa."""
    calc = db.query(MonthlySalaryCalculation).filter(
        MonthlySalaryCalculation.id == calculation_id
    ).first()
    if not calc:
        raise ValueError("Hisob-kitob topilmadi")

    if calc.status == "paid":
        raise ValueError("Ushbu xodimga oylik allaqachon to'langan!")

    emp = calc.employee
    cash_reg = db.query(CashRegister).filter(CashRegister.id == register_id).first()
    if not cash_reg:
        raise ValueError("Kassa hisobi topilmadi")

    # Generate Cash Transaction (Chiqim)
    desc = f"Ish haqi to'lovi ({calc.year_month}): {emp.full_name}"
    if notes:
        desc += f" ({notes})"

    tx = CashTransaction(
        register_id=register_id,
        type="chiqim",
        source_type="other",
        amount=payment_amount,
        currency=cash_reg.currency,
        category="Ishchilar oyligi / Avans",
        date=date.today(),
        description=desc,
        status="Tasdiqlandi"
    )
    db.add(tx)
    
    # Update Cash Register balance
    cash_reg.balance = (cash_reg.balance or 0.0) - payment_amount
    db.flush()

    # Update calculation status
    calc.status = "paid"
    calc.paid_at = datetime.utcnow()
    calc.paid_by = current_user
    calc.cash_transaction_id = tx.id
    if notes:
        calc.notes = (calc.notes or "") + f" [To'lov: {notes}]"

    db.add(AuditLog(
        username=current_user,
        action="PAYMENT",
        module="Ish haqi / Kassa",
        entity_id=str(calc.id),
        details=f"Oylik to'landi: {emp.full_name} ({calc.year_month}) -> {payment_amount:,.0f} {cash_reg.currency} (Kassa #{register_id})"
    ))
    db.commit()
    db.refresh(calc)
    return calc

def generate_payroll_excel(db: Session, year_month: str) -> io.BytesIO:
    """Export monthly payroll table to a beautifully formatted Excel sheet."""
    summary = get_payroll_summary(db, year_month)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ish haqi {year_month}"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_title = Font(name="Arial", size=15, bold=True, color="1E3A8A")
    font_sub = Font(name="Arial", size=10, bold=True, color="64748B")
    font_th = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_td = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Title Block
    ws.merge_cells("A1:K1")
    ws["A1"] = f"KAFEL ZAVODI — OYLIK ISH HAQI VEDOMOSTI ({year_month})"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Jami hisoblangan: {summary['total_payroll']:,.0f} UZS  |  Fiks: {summary['total_fixed']:,.0f} UZS  |  Ishbay: {summary['total_piecework']:,.0f} UZS  |  To'langan: {summary['total_paid']:,.0f} UZS"
    ws["A2"].font = font_sub
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        "№", "F.I.SH. (Xodim)", "Lavozimi", "Turi", 
        "Fiks oylik (UZS)", "Ish kuni", "Kelmadi", "Ushlanma (UZS)",
        "Ishbay jami (UZS)", "Qo'shimcha / Avans", "Jami to'lov (UZS)", "Holati"
    ]
    
    ws.row_dimensions[4].height = 24
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_th
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Rows
    current_row = 5
    for idx, c in enumerate(summary["calculations"], start=1):
        emp_type_label = "Fiksalangan" if c["employee_type"] == "fixed" else "Ishbay"
        status_label = "To'langan" if c["status"] == "paid" else ("Tasdiqlangan" if c["status"] == "finalized" else "Qoralama")
        bonus_adv_str = f"+{c['bonus_amount']:,.0f} / -{c['advance_paid']:,.0f}" if (c['bonus_amount'] or c['advance_paid']) else "0"

        row_vals = [
            idx,
            c["full_name"],
            c["position"],
            emp_type_label,
            c["base_salary"] if c["employee_type"] == "fixed" else 0,
            c["standard_days"] if c["employee_type"] == "fixed" else "-",
            c["absent_days"] if c["employee_type"] == "fixed" else "-",
            c["deduction_amount"] if c["employee_type"] == "fixed" else 0,
            c["piecework_total"] if c["employee_type"] == "piecework" else 0,
            bonus_adv_str,
            c["final_amount"],
            status_label
        ]

        ws.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_td
            cell.border = thin_border
            if current_row % 2 == 0:
                cell.fill = fill_zebra
                
            if col_idx in (5, 8, 9, 11) and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (1, 6, 7, 10, 12):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # Total Row
    ws.row_dimensions[current_row].height = 24
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    total_cell = ws.cell(row=current_row, column=1, value="JAMI OYLIK ISH HAQI:")
    total_cell.font = font_bold
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.fill = fill_total
    
    for c_i in range(1, 11):
        ws.cell(row=current_row, column=c_i).fill = fill_total
        ws.cell(row=current_row, column=c_i).border = thin_border

    sum_cell = ws.cell(row=current_row, column=11, value=summary["total_payroll"])
    sum_cell.font = font_bold
    sum_cell.number_format = '#,##0'
    sum_cell.alignment = Alignment(horizontal="right", vertical="center")
    sum_cell.fill = fill_total
    sum_cell.border = thin_border

    ws.cell(row=current_row, column=12, value="").fill = fill_total
    ws.cell(row=current_row, column=12).border = thin_border

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
