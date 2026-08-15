from datetime import date, datetime
import io
from typing import Dict, Any, List
from sqlalchemy import extract
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from backend.models import (
    Sale, Purchase, ProductionOrder, CashTransaction, MonthClosing,
    StockItem, MDMMaterial, MDMCounterparty, Warehouse
)
from backend.services.currency_service import convert_amount
from backend.services.cost_allocation_service import calculate_monthly_production_cost_allocation

def get_pnl_report(db: Session, year_month: str) -> Dict[str, Any]:
    year, month = map(int, year_month.split("-"))

    # 1. Total Revenue in USD
    sales = db.query(Sale).filter(
        extract('year', Sale.date) == year,
        extract('month', Sale.date) == month,
        Sale.status == "Tasdiqlandi"
    ).all()

    total_revenue_usd = 0.0
    for s in sales:
        total_revenue_usd += convert_amount(s.total_amount, s.currency, "USD", s.date, db)

    # 2. Production Cost Allocation for the 5 Lines
    alloc = calculate_monthly_production_cost_allocation(db, year_month)
    direct_materials_cogs = alloc["total_direct_materials_cost_usd"]
    indirect_expenses_cogs = alloc["total_indirect_expenses_usd"]
    total_cogs_usd = direct_materials_cogs + indirect_expenses_cogs

    gross_profit_usd = total_revenue_usd - total_cogs_usd
    admin_expenses_usd = alloc["total_admin_expenses_usd"]
    net_profit_usd = gross_profit_usd - admin_expenses_usd

    # Check if month is closed
    month_close = db.query(MonthClosing).filter(MonthClosing.year_month == year_month).first()
    is_closed = month_close.is_closed if month_close else False

    return {
        "year_month": year_month,
        "currency": "USD",
        "revenue_usd": round(total_revenue_usd, 2),
        "cogs_direct_materials_usd": round(direct_materials_cogs, 2),
        "cogs_indirect_expenses_usd": round(indirect_expenses_cogs, 2),
        "total_cogs_usd": round(total_cogs_usd, 2),
        "gross_profit_usd": round(gross_profit_usd, 2),
        "admin_expenses_usd": round(admin_expenses_usd, 2),
        "net_profit_usd": round(net_profit_usd, 2),
        "is_closed": is_closed,
        "total_factory_volume_m2": alloc["total_factory_volume_m2"],
        "line_breakdown": alloc["lines"]
    }

def get_cash_flow_report(db: Session, year_month: str) -> Dict[str, Any]:
    year, month = map(int, year_month.split("-"))

    txs = db.query(CashTransaction).filter(
        extract('year', CashTransaction.date) == year,
        extract('month', CashTransaction.date) == month
    ).all()

    inflows_usd = 0.0
    outflows_usd = 0.0
    categories: Dict[str, Dict[str, float]] = {}

    for tx in txs:
        usd_amt = convert_amount(tx.amount, tx.currency, "USD", tx.date, db)
        cat = tx.category
        if cat not in categories:
            categories[cat] = {"inflow": 0.0, "outflow": 0.0}

        if tx.type == "kirim":
            inflows_usd += usd_amt
            categories[cat]["inflow"] += usd_amt
        else: # chiqim
            outflows_usd += usd_amt
            categories[cat]["outflow"] += usd_amt

    breakdown = []
    for cat, val in categories.items():
        breakdown.append({
            "category": cat,
            "inflow_usd": round(val["inflow"], 2),
            "outflow_usd": round(val["outflow"], 2),
            "net_usd": round(val["inflow"] - val["outflow"], 2)
        })

    net_cash_flow_usd = inflows_usd - outflows_usd

    return {
        "year_month": year_month,
        "total_inflows_usd": round(inflows_usd, 2),
        "total_outflows_usd": round(outflows_usd, 2),
        "net_cash_flow_usd": round(net_cash_flow_usd, 2),
        "breakdown_by_category": breakdown
    }

def generate_stock_excel(db: Session, warehouse_id: int = None) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ombor qoldiqlari"

    # Styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin', color="D1D5DB"),
        right=Side(style='thin', color="D1D5DB"),
        top=Side(style='thin', color="D1D5DB"),
        bottom=Side(style='thin', color="D1D5DB")
    )

    headers = [
        "Sklad (Ombor)", "Mahsulot Kodi", "Mahsulot Nomi", "Kategoriyasi",
        "O'lchov Birligi", "Miqdor (Qoldiq)", "AVG Narxi (USD)", "AVG Narxi (UZS)",
        "Jami Qiymat (USD)", "Jami Qiymat (UZS)"
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    query = db.query(StockItem)
    if warehouse_id:
        query = query.filter(StockItem.warehouse_id == warehouse_id)
    items = query.all()

    for item in items:
        total_usd = item.quantity * item.avg_cost_usd
        total_uzs = item.quantity * item.avg_cost_uzs
        ws.append([
            item.warehouse.name if item.warehouse else "-",
            item.material.code if item.material else "-",
            item.material.name if item.material else "-",
            item.material.category if item.material else "-",
            item.material.unit if item.material else "-",
            round(item.quantity, 2),
            round(item.avg_cost_usd, 4),
            round(item.avg_cost_uzs, 2),
            round(total_usd, 2),
            round(total_uzs, 2)
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_mdm_excel(db: Session, entity_type: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = entity_type.capitalize()

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    if entity_type == "materials":
        headers = ["Kodi", "Nomi", "Kategoriyasi", "Birligi", "Min Qoldiq", "Joriy AVG ($)", "Holati"]
        ws.append(headers)
        materials = db.query(MDMMaterial).all()
        for m in materials:
            ws.append([
                m.code, m.name, m.category, m.unit, m.min_stock,
                round(m.current_avg_price_usd, 4),
                "Arxivlangan" if m.is_archived else "Faol"
            ])
    else: # counterparties (clients / suppliers)
        headers = ["Kodi", "Nomi", "Turi", "Rezident", "Viloyat", "Telefon", "Boshlang'ich Balans ($)", "Joriy Balans ($)", "Joriy Balans (UZS)", "Holati"]
        ws.append(headers)
        cps = db.query(MDMCounterparty)
        if entity_type == "clients":
            cps = cps.filter(MDMCounterparty.type == "client")
        elif entity_type == "suppliers":
            cps = cps.filter(MDMCounterparty.type == "supplier")
        for cp in cps.all():
            ws.append([
                cp.code, cp.name, cp.type, "Ha" if cp.is_resident else "Yo'q",
                cp.region, cp.phone or "-", cp.initial_balance_usd,
                round(cp.current_balance_usd, 2), round(cp.current_balance_uzs, 2),
                "Arxivlangan" if cp.is_archived else "Faol"
            ])

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
